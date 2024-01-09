"""
Макросейсмика, 2020,
главный модуль
"""
# pmain_proc

import numpy as np
import scipy
# import scipy.stats as st
# from scipy.stats import skew
# from scipy.stats import kurtosis

from pinp_proc import *

import pinp_struct

from pnumpy import *
import math
import numba
# from scipy.optimize import minimize
import copy
import openpyxl
# from psort import *
# from pfit import *
from tkinter import messagebox as mb
from ptkinter_menu_proc import *
from dataclasses import dataclass
from sklearn.linear_model import LinearRegression

log_file_name: str
gran = 0.001  # 0.001% изменения функции по абсолютной величине
# для analyze_diff, порог для max_var/max_fun, возможные уровни срабатывания 1, 10, 20, 100, 400

def get_data_ini2():
    (the_dict, the_arr) = pinp_struct.get_dat_struct(pinp_struct.curr_nstruct)
    grav_ = the_arr[:, 1]
    sd_ = the_arr[:, 2]
    tilt_x = the_arr[:, 3]
    tilt_y = the_arr[:, 4]
    temp_ = the_arr[:, 5]
    tide_ = the_arr[:, 6]
    dur_ = the_arr[:, 7]
    rej_ = the_arr[:, 8]
    time_ = the_arr[:, 9]
    date_time_ = the_arr[:, 10]
    date_ = the_arr[:, 11]
    llen = date_time_.size
    xn_ = np.arange(0.0, llen, 1.0)
    # for i in range(30):
    #     print(grav_[i],' ',sd_[i],' ', tilt_x[i],' ', tilt_y[i],' ', temp_[i],' ', tide_[i],' ', dur_[i],' ',\
    #           rej_[i],' ', time_[i],' ', date_time_[i],' ',xn_[i])
    return grav_, sd_, tilt_x, tilt_y, temp_, tide_, dur_, rej_, date_time_, xn_, time_, date_


@dataclass
class DescrStat:
    # работа с 1 мерными массивами NumPy
    name: str =''
    n: int = None # длина массива - x.size
    min_: float = None
    max_: float = None
    range_: float = None
    mean_: float = None
    st_mean_: float = None # стандартная ошибка среднего
    median_: float = None
    quant_025: float = None # 0.25 квантиль, 1 квартиль
    quant_050: float = None # 0.50 квантиль, 2 квартиль
    quant_075: float = None # 0.75 квантиль, 3 квартиль
    std_vib: float = None # стандартное отклонение выборки, ddof=1
    std_gen: float = None # стандартное отклонение ген.совокуп, ddof=0
    var_vib: float = None # дисперсия выборки, ddof=1
    var_gen: float = None # дисперсия ген.совокуп, ddof=0
    skew_biasT: float = None# Асимметрия скорректированное за статист. смещение
    skew_biasF: float = None # Асимметрия НЕскорректированное за статист. смещение
    kurt_biasT: float = None# Эксцесс скорректированн за статист. смещение
    kurt_biasF: float = None # Эксцесс НЕскорректированн за статист. смещение
    # Как рассчитать доверительные интервалы в Python, 17 авг. 2022 г.
    # https://www.codecamp.ru/blog/confidence-intervals-python/
    # считаю с использованием t-распеределения
    min95_confid: float = None # минимальное значение 95% доверительного интервала для среднего
    max95_confid: float = None # минимальное значение 95% доверительного интервала для среднего


def calc_descr_stat(x, name='', is_view=False):
    '''
    Расчет описательной статистики, результирующие параметры аналогичны Excel
    '''
    ptp_ = mean_ = st_mean_ = median_ = quant_025 = quant_050 = quant_075 = std_vib = std_gen = None
    var_vib = var_gen = min95_confid = max95_confid = None
    s = ''
    name_ = name
    n_ = x.size
    min_ = float(np.min(x))
    max_ = float(np.max(x))
    equal_data = abs(min_ - max_) < 1e-38
    if not equal_data:
        ptp_ = float(np.ptp(x))
        mean_ = float(np.mean(x))
        st_mean_ = float(scipy.stats.sem(x)) # стандартная ошибка среднего
        median_ = float(np.median(x))
        quant_025 = float(np.quantile(x, 0.25))  # linear распределение по умолчанию
        quant_050 = float(np.quantile(x, 0.50))  # linear распределение по умолчанию
        quant_075 = float(np.quantile(x, 0.75))  # linear распределение по умолчанию
        std_vib = float(np.std(x,ddof=1))       # стандартное отклонение выборки, ddof=1
        std_gen = float(np.std(x,ddof=0))       # стандартное отклонение ген.совокуп, ddof=0
        var_vib = float(np.var(x,ddof=1))       # дисперсия выборки, ddof=1
        var_gen = float(np.var(x,ddof=0))       # дисперсия ген.совокуп, ddof=0

        # skew_biasT = scipy.stats.skew(x, bias=True) # Асимметрия скорректированное за статист. смещение
        # skew_biasF = float(scipy.stats.skew(x, bias=False)) # Асимметрия НЕскорректированное за статист. смещение
        # kurt_biasT: float = float(kurtosis(x, bias=True)) # Эксцесс скорректированн за статист. смещение
        # kurt_biasF: float = float(kurtosis(x, bias=False)) # Эксцесс НЕскорректированн за статист. смещение

        ttt = scipy.stats.t.interval(confidence=0.95, df=n_-1, loc=mean_, scale=st_mean_)
        min95_confid = ttt[0] # минимальное значение 95% доверительного интервала для среднего
        max95_confid = ttt[1] # максимальное значение 95% доверительного интервала для среднего

    the_descr_stat = DescrStat(name_, n_, min_, max_, ptp_, mean_, st_mean_, median_, quant_025, quant_050,
                               quant_075, std_vib, std_gen, var_vib, var_gen, min95_confid, max95_confid)
    s +=f'Описательная статистика для {name_} \n'
    s+= f'Отсчетов = {n_} \n'
    s+= f'min      = {min_} \n'
    s+= f'max      = {max_} \n'
    if not equal_data:
        s+= f'max-min    = {max_-min_} \n'
        s+= f'range(ptp) = {ptp_} \n'
        s+= f'mean    = {mean_} \n'
        s+= f'стандартная ошибка среднего = {st_mean_} \n'
        s+= f'median  = {median_} \n'
        s+= f'quantile 0.25 (linear) = {quant_025} \n'
        s+= f'quantile 0.50 (linear) = {quant_050} \n'
        s+= f'quantile 0.75 (linear) = {quant_075} \n'

        s+= f'стандартное отклонение выборки = {std_vib} \n'
        s+= f'стандартное ген.совокупности   = {std_gen} \n'
        s+= f'дисперсия выборки              = {var_vib} \n'
        s+= f'дисперсия ген.совокупности     = {var_gen} \n'

        # s+= f'Асимметрия скорректированное за статист. смещение = {skew_biasT} \n'
        # s+= f'Асимметрия НЕскорректированное за статист. смещение = {skew_biasF} \n'
        # s+= f'Эксцесс скорректированн за статист. смещение = {kurt_biasT} \n'
        # s+= f'Эксцесс Нескорректированн за статист. смещение = {kurt_biasF} \n'
        s+= f'Минимальное значение 95% доверительного интервала для среднего = {min95_confid} \n'
        s+= f'Максимальное значение 95% доверительного интервала для среднего = {max95_confid} \n \n'

    if is_view:
        print(s)
    return the_descr_stat, s

def calc_stat():
    #   0      1     2       3       4      5      6      7     8         9    10     11
    #  grav_, sd_, tilt_x, tilt_y, temp_, tide_, dur_, rej_, date_time_, xn_, time_, date_
    datas = get_data_ini2()
    s1=''
    for i in range(len(full_grav_data_name)-1): # full_grav_data_name
        print(i, full_grav_data_name[i],'   ',  type(datas[i]),'   ',datas[i].size)
        (the_descr_stat, s) = calc_descr_stat(datas[i], full_grav_data_name[i], False)
        s1+=s
    return s1

def linear_trend(x,y, is_view=False)->(float, float, float, np.ndarray, np.ndarray):
    x_ = x.reshape((-1, 1))
    # 25/04/2019 Линейная регрессия на Python: объясняем на пальцах https://proglib.io/p/linear-regression
    model = LinearRegression()
    model.fit(x_, y)
    r_sq = model.score(x_, y)
    if is_view: print('coefficient of determination:', r_sq)
    # y = np.interp(xn_, xp, fp)  # Расчет в точках интерполяции
    intercept: 5.633333333333329
    if is_view: print('Наклон:', model.coef_)
    y_pred = model.predict(x_)
    ost_anom = y - y_pred
    return r_sq, model.intercept_, model.coef_, y_pred, ost_anom

def moving_averages(data:np.ndarray, kernel_size:int)->np.ndarray:
    '''
    02.06.2020 Сглаживание данных с помощью скользящего среднего с помощью NumPy
    https://danielmuellerkomorowska.com/2020/06/02/smoothing-data-by-rolling-average-with-numpy/
    '''
    kernel = np.ones(kernel_size) / kernel_size
    data_convolved = np.convolve(data, kernel, mode='same')
    return data_convolved

# def work_with_data(is_view1: bool) -> (bool, int, float, float, float, float, float, list):
#     """
#     Подготовка к работе и минимизация
#     bool
#     """
#     row: int; i: int;
#
#     if is_view1: print('work_with_data')
#     (the_dict, the_arr) = pinp_struct.get_dat_struct(pinp_struct.curr_nstruct)
#     row = the_dict["npoint"]
#     lat_arr = the_arr[:, 0]  # Lat
#     # print('lat_arr ', np.size(lat_arr)); print(lat_arr)
#     lon_arr = the_arr[:, 1]  # Lat
#     # print('lon_arr ', np.size(lon_arr)); print(lon_arr)
#     h_arr = the_arr[:, 2]/1000  # Alt переводим в км
#     # print('h_arr ', np.size(h_arr)); print(h_arr)
#     i_fact_arr = the_arr[:, 3]
#     npoint = np.size(i_fact_arr)
#     eqval = (row == npoint)
#     # print('i_fact_arr ', np.size(i_fact_arr)); print(i_fact_arr)
#     # if not eqval:
#     #     # заглушка для данных
#     #     num, lat_, lon_, dep_, mag_, fun_ = -13, -13, -13, -13, -13, -13,
#     # else:
#     #     (num, lat_, lon_, dep_, mag_, fun_, res_list_) = minimize_func(npoint, lat_arr, lon_arr, h_arr, i_fact_arr)
#     # return eqval, num, lat_, lon_, dep_, mag_, fun_, res_list_
#     # view_2d_array(dat, nrow: int, ncol: int, test_info='Просмотр в work_with_data '):
#
#
# def calc_diff(iii: int, x0: float, x1: float, x2: float, x3: float, f: float, result_list: list) -> \
#                                                         (float, float, float, float, float, float):
#                                                         # dx0    dx1   dx2    dx3,   df,    gran
#     if iii == 0:
#         return math.nan, math.nan, math.nan, math.nan, math.nan, math.nan
#     else:
#         n = 100
#         dat = result_list[iii-1]
#         dx0 = ((x0 - dat[0])/dat[0])*n  # lat_ = x0[0]
#         dx1 = ((x1 - dat[1])/dat[1])*n  # lon_ = x0[1]
#         dx2 = ((x2 - dat[2])/dat[2])*n  # dep_ = x0[2]
#         dx3 = ((x3 - dat[3])/dat[3])*n  # mag_ = x0[3]
#         df  = ((f  - dat[4])/dat[4])*n  # f = x0[4]
#         dgran_ = get_diffgran(dx0, dx1, dx2, dx3, df)
#         return (dx0, dx1, dx2, dx3, df, dgran_)
#
#
# def get_diffgran(dx0: float, dx1: float, dx2: float, dx3: float, df: float) -> float:
#     # max_dvar = abs(max(dx0, dx1, dx2, dx3))
#     max_dvar = max(abs(dx0), abs(dx1), abs(dx2), abs(dx3))
#     max_dfun = abs(df)
#     # max_dvar/max_dfun - числа не более 1, 10, 40 - как бы "линейная" зависимость
#     # max_dfun/max_dvar - как бы парабола, ветви вверх, аппроксимируем, находим индекс минимума
#     # как бы 2-3 минимума, но они после больших значений, так что реагировать будут на первый
#     return max_dvar/max_dfun
#
#
# def analyze_diff(iii: int, result_list: list) -> bool:
#     """
#     Определяет как быстро меняется самая вариабельная переменная относительно изменения функции
#     Если более gran, то return True / иначе return False
#     """
#     global gran
#     if iii != 0:   # 5    6     7   8     9
#         dat = result_list[iii]
#         # max_var = max(abs(dat[5]), abs(dat[6]), abs(dat[7]), abs(dat[8]))
#         max_fun = abs(dat[9])
#         # d = max_var/max_fun  # как быстро меняется самая вариабельная переменная относительно изменения функции
#         d = max_fun  # как быстро меняется целевая функция
#         # print('max_var/max_fun = %7.5f' %  d)
#         if d > gran:  # если  - числа не более 1, 10, 40 - как бы "линейная" зависимость
#             # print('d > gran')
#             return True
#         else:
#             return False
#
#
# def output_txt_res(fname: str, result_list: list) -> None:
#     """
#     Неполный вариант по выводим данным, полный см. output_xls_res
#     Вывод в txt файл итераций минимизации:
#     номер, lat, lon, dep, mag, fun
#     """
#     fff = open(fname,  mode='w')  # открытие log-файла
#     llen = len(result_list)
#     fff.write('{0:6s} {1:9s} {2:9s} {3:9s} {4:9s} {5:11s} {6:9s} {7:9s} {8:9s} {9:9s} {10:11s} {11:11s}\n'.format('N', 'Lat', 'Lon', 'Dep', 'Mag', 'Fun', 'dLat,%', 'dLon,%', 'dDep,%', 'dMag,%', 'dFun,%', 'max_var,%/dFun,%'))
#     for i in range(llen):
#         d = result_list[i]
#         fff.write('{0:5d} {1:9.5f} {2:9.5f} {3:9.5f} {4:9.5f} {5:11.5f} {6:9.5f} {7:9.5f} {8:9.5f} {9:9.5f}  {10:11.5f} {11:11.5f}\n'.format(i, d[0], d[1], d[2], d[3], d[4], d[5], d[6], d[7], d[8], d[9], d[10]))
#     fff.close()
#
#
# def output_xls_res(fname: str, result_list: list) -> None:
#     """
#     Вывод в xlsx файл итераций минимизации:
#     номер, lat, lon, dep, mag, fun, dLat %, dLon %, dDep %, dMag %, dFun %
#     """
#     # Чтение и запись в файл Excel с использованием модуля Python openpyxl
#     # https://pythononline.ru/question/chtenie-i-zapis-v-fayl-excel-s-ispolzovaniem-modulya-python-openpyxl
#     my_wb = openpyxl.Workbook()
#     my_sheet = my_wb.active
#     my_sheet.title = "Результаты"
#     llen = len(result_list)
#     my_sheet.cell(row=1, column= 1).value = 'i'
#     my_sheet.cell(row=1, column= 2).value = 'Lat'
#     my_sheet.cell(row=1, column= 3).value = 'Lon'
#     my_sheet.cell(row=1, column= 4).value = 'Dep'
#     my_sheet.cell(row=1, column= 5).value = 'Mag'
#     my_sheet.cell(row=1, column= 6).value = 'Fun'
#     my_sheet.cell(row=1, column= 7).value = 'dLat, %'
#     my_sheet.cell(row=1, column= 8).value = 'dLon, %'
#     my_sheet.cell(row=1, column= 9).value = 'dDep, %'
#     my_sheet.cell(row=1, column=10).value = 'dMag, %'
#     my_sheet.cell(row=1, column=11).value = 'dFun, %'
#     my_sheet.cell(row=1, column=12).value = 'max_var,%/dFun,%'
#     for i in range(llen):
#         d = result_list[i]
#         rrow = i + 2
#         my_sheet.cell(row=rrow, column=1).value = i  # i
#         my_sheet.cell(row=rrow, column=2).value = d[0]  # Lat
#         my_sheet.cell(row=rrow, column=3).value = d[1]  # Lon
#         my_sheet.cell(row=rrow, column=4).value = d[2]  # Dep
#         my_sheet.cell(row=rrow, column=5).value = d[3]  # Mag
#         my_sheet.cell(row=rrow, column=6).value = d[4]  # Fun
#         if rrow > 2:
#             my_sheet.cell(row=rrow, column= 7).value = d[5]  # dLat
#             my_sheet.cell(row=rrow, column= 8).value = d[6]  # dLon
#             my_sheet.cell(row=rrow, column= 9).value = d[7]  # dDep
#             my_sheet.cell(row=rrow, column=10).value = d[8]  # dMag
#             my_sheet.cell(row=rrow, column=11).value = d[9]  # dFun
#             my_sheet.cell(row=rrow, column=12).value = d[10]  # max(dvar)/dFun
#
#     my_wb.save(fname)

