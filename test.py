import datetime

import numpy as np
import openpyxl
import pfile
import pathlib
import os
from scipy import signal
import random
from copy import deepcopy
import matplotlib.pyplot as plt
from scipy.stats import skew
from scipy.stats import kurtosis



# date_time_str = '2018-06-29 08:15:27.243860'
# date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S.%f')
# print('Дата:', date_time_obj.date())
# print('Время:', date_time_obj.time())
# print('Дата и время:', date_time_obj, '    ', type(date_time_obj ))

# date_time_str = '03.04.17 9:33'
# date_time_obj = datetime.datetime.strptime(date_time_str, '%d.%m.%y %H:%M')
# print('Дата:', date_time_obj.date())
# print('Время:', date_time_obj.time())
# print('Дата и время:', date_time_obj, '    ', type(date_time_obj ))


# ## opening the previously created xlsx file using 'load_workbook()' method
# xlsx = openpyxl.load_workbook('J:/Work-Lang2/Python/PyCharm/GTimeSeries/dat/Sample.xlsx')
# ## getting the sheet to active
# sheet = xlsx.active
# ## getting the reference of the cells which we want to get the data from
# name = sheet['A1']
# tag = sheet.cell(row = 1, column = 2) # 2 in cell
# ## printing the values of cells
# print(name.value)
# print(tag.value)

# s ='18.02.17 9:52'
# s2 = s.split(sep=' ', maxsplit=2)  # print_string(part_lines)
# print(s2[0].strip())
# print(s2[1].strip())

# for i in range(2,10+1):
#     print(i)

# from datetime import datetime
# ll =['03.04.17 9:33', '03.04.17 9:34', '03.04.17 9:35', '03.04.17 9:36', '03.04.17 9:37', '03.04.17 9:38', '03.04.17 9:39']
# print(ll)
# kk=[]
# for i in range(len(ll)):
#     dd = datetime.strptime(ll[i], '%d.%m.%y %H:%M')
#     kk+=[dd.timestamp()]
# print(kk)
#
# zz =[]
# for i in range(len(kk)):
#     hhh = datetime.fromtimestamp(kk[i], tz=None)
#     zz+=[hhh]
# print(zz)

# x = np.array(kk)
# min_: float = float(np.min(x))
# max_: float = float(np.max(x))
# ptp_: float = float(np.ptp(x))
# mean_: float = float(np.mean(x))
#
# print(min_)
# print(max_)
# print(ptp_)
# print(mean_)


# now = datetime.datetime.now()
# then = datetime.datetime(2023, 4, 9)
# # Кол-во времени между датами.
# delta = now - then
# print(delta)
# print(delta.total_seconds())
# print(delta.total_seconds()/60,'min')
# print(delta.total_seconds()/60/60,'hour')

# n = 600
# t = np.arange(0,n,1)
# sig = deepcopy(t)
# for i in range(n):
#     sig[i] = random.gauss(0, 0.5)
# sig_t = sig + t*0.001
# sig_d = signal.detrend(sig_t, type='linear') # +sig[0]
# plt.plot(t, sig, "x", label ='sig')
# plt.plot(t, sig_t, label ='sig_t')
# plt.plot(t, sig_d, label ='sig_d')
# plt.legend(loc='upper right'); plt.grid(); plt.show()

# fn = r'j:/Data.txt'
# n = pfile.text_file_num_lines(fn)
# a = np.zeros(n)
# print('\n',fn)
# f = open(fn, 'r')
# for i in range(n):
#     s = f.readline()
#     a[i] = float(s)
# print(a)
# print(type(a))
# skew_biasT: float = float(skew(a, bias=True)) # Асимметрия скорректированное за статист. смещение
# print(skew_biasT)
# skew_biasF: float = float(skew(a, bias=False)) # Асимметрия НЕскорректированное за статист. смещение
# print(skew_biasF)
# kurt_biasT: float = float(kurtosis(a, bias=True)) # Эксцесс скорректированн за статист. смещение
# print(kurt_biasT)
# kurt_biasF: float = float(kurtosis(a, bias=False)) # Эксцесс НЕскорректированн за статист. смещение
# print(kurt_biasF)

# a = 10.0
# b = 2.0
# print(type(a**b))
# print(a**b)

# https://pythonstart.ru/list/clear-python?ysclid=lr6ji8tclb389253375
# lst = [{1, 2}, ('a'), ['1.1', '2.2']]
# lst.clear()
# print(lst,' ',lst == [])
# lst = [{1, 2}, ('a'), ['1.1', '2.2']]
# lst =[]
# print(lst,' ',lst == [])

l1= [0,1,2,3,]
l2= [0,1,2,3,]
plt.plot(l1,l2)
plt.show()