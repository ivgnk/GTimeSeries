'''
# Ввод данных из xlsx - файла
#
# (C) 2024 Ivan Genik, Perm, Russia
# Released under GNU Public License (GPL)
# email igenik@rambler.ru

import openpyxl
import numpy as np

'''
import numpy as np
import openpyxl
import scipy


#################################### Import Data

def the_xls_importdat(xls_file_name: str, is_view: bool) -> np.ndarray:
    """
    ВВод данных из excel-файла
    образец файл: Dat\точки_ввод.xlsx
    """
    nrow: int; ncol: int
#  2020 Чтение и запись в файл Excel с использованием модуля Python openpyxl
#  Источник: https://pythononline.ru/question/chtenie-i-zapis-v-fayl-excel-s-ispolzovaniem-modulya-python-openpyxl
    my_wb = openpyxl.load_workbook(xls_file_name, data_only=True)
    my_sheet = my_wb.active
    my_sheet_title = my_sheet.title
    if is_view:
        print("My sheet title: " + my_sheet_title)
# Openpyxl - Как найти количество строк с данными в xlsx
# Источник: https://question-it.com/questions/611944/openpyxl-kak-najti-kolichestvo-strok-s-dannymi-v-xlsx
# пустые столбцы и строки определяет плохо
#     if is_view: print('my_sheet.max_row = ', my_sheet.max_row)
#     if is_view: print('my_sheet.max_column = ',my_sheet.max_column)
    if is_view: view_excel_sheet(my_sheet, 40)
    # удаление лишних столбцов и создание нового с точным датой и временем
    # Подготовка массива NumPy
    # https://coderoad.ru/53659234/Как-записать-строку-в-массив-numpy
    nrow = my_sheet.max_row;    ncol = my_sheet.max_column
    # if is_view:
    #     print('nrow=',nrow)
    #     print('ncol=',ncol)
    #     input('press any key to continue')


    numpy_arr = np.zeros((nrow-1, ncol), dtype=object)  # первую строку заголовков не вводим
    # if is_view:   view_2d_array(numpy_arr, nrow-1, ncol, '2d массив NumPy создан')
    #  извлечение из my_sheet
    for i in range(nrow):
        if i != 0:  # первую строку заголовков не вводим
            for j in range(ncol):
                numpy_arr[i-1, j] = my_sheet.cell(row=i+1, column=j+1).value           # первую строку заголовков не вводим
    # if the_view:  view_2d_array(numpy_arr, nrow - 1, ncol, '2d массив NumPy заполнен')
    # if is_view:  view_2d_array(numpy_arr, nrow - 1, ncol, '2d массив NumPy заполнен')

    return numpy_arr
# --------------- def the_xls_import()


def view_excel_sheet(excel_sheet, n_view_row=-1) -> None:
    print('2d лист excel')
    nrow: int = excel_sheet.max_row
    ncol: int = excel_sheet.max_column
    print('excel_sheet.max_row = ', nrow)
    print('excel_sheet.max_column = ', ncol)

    n = n_view_row if n_view_row>0 else nrow
    shift_str = ''.join([' ']*9) # repeat_chr(' ',9)
    for i in range(n):
        # Использование Python и Excel для обработки и анализа данных. Часть 2: библиотеки для работы с данными
        # https://habr.com/ru/company/otus/blog/331998/
        if i != 0: # номера значений
            print(format(i, '7d'), end='')
        else:
            print(shift_str, end='')
        for j in range(ncol):
            s = shift_str if j == 0 else ''  # сдвиг из-за номера значений
            s = s + str(excel_sheet.cell(row=i+1, column=j+1).value)
            lens = len(s)
            if lens<=9: the_s = '9s'
            else: the_s = str(lens)+'s'
            print(format(s,the_s), end='  ')
        print()

#################################### Descriptive Ststistics
