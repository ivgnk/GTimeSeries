# -------------------------------------------------------------------
# -------------------------------------------------------------------
# Ввод из данных с координатами и данными точек
# xlsx - файла
# txt - файла
#
# (C) 2023 Ivan Genik, Perm, Russia
# Released under GNU Public License (GPL)
# email igenik@rambler.ru
# -------------------------------------------------------------------
# Литература
# 2010 Обработка Excel файлов с использованием Python
# https://habr.com/ru/post/99923/
# 2014 # Интеграция MS Excel и Python
# https://habr.com/ru/post/232291/

# 2016 how to open xlsx file with python 3 (xlrd)
# https://stackoverflow.com/questions/37434227/how-to-open-xlsx-file-with-python-3

# Использование Python и Excel для обработки и анализа данных. Часть 1: импорт данных и настройка среды (pandas)
# https://habr.com/ru/company/otus/blog/331746/
# 2017 Использование Python и Excel для обработки и анализа данных. Часть 2: библиотеки для работы с данными
# https://habr.com/ru/company/otus/blog/331998/

# 2018 Работа с файлами Excel в Python
# https://tokmakov.msk.ru/blog/item/71
# 2018 Работа с Excel в Питоне
# http://amcher.ru/2018/11/16/rabota-s-excel-v-pitone/

# 2019 Python  создание и запись в файл Excel с помощью модуля Xlsxwriter (xlsxwriter)
# http://espressocode.top/python-create-and-write-on-excel-file-using-xlsxwriter-module/
# How to write a complex numpy array to excel using openpyxl
# https://stackoverflow.com/questions/55875728/how-to-write-a-complex-numpy-array-to-excel-using-openpyxl


# 2020 Работа с Excel из Python: Полное руководство с примерами (pandas, openpyxl)
# https://codecamp.ru/blog/python-excel-tutorial/
# 2020 Чтение и запись файлов Excel (XLSX) в Python (pandas)
# https://pythonru.com/uroki/chtenie-i-zapis-fajlov-excel-xlsx-v-python
# 2020 Пишем файл Excel из Python (XlsxWriter)
# https://tirinox.ru/write-excel-python/ {Python - site}
# 2020 Чтение и запись в файл Excel с использованием модуля Python openpyxl
# https://pythononline.ru/question/chtenie-i-zapis-v-fayl-excel-s-ispolzovaniem-modulya-python-openpyxl {Python - site}
#

# https://openpyxl.readthedocs.io/en/stable/tutorial.html

# Excel и Python. Основы взаимодействия (pandas)
# http://www.excelguide.ru/2019/12/excel-python-basics.html

# NUMPY
# Python 3: обработка массивов numpy и экспорт через openpyxl
# https://fooobar.com/questions/12725842/python-3-handling-numpy-arrays-and-export-via-openpyxl

# 2018 Как записать строку в массив numpy?
# https://coderoad.ru/53659234/Как-записать-строку-в-массив-numpy

# -------------------------------------------------------------------
# pinp_proc

import sys
import openpyxl
import numpy as np
import random
import pathlib

import pfile
from pfile import *
from pstring import *
from pinp_struct import *
from pnumpy import *
from tkinter import messagebox as mb
import pmain_proc

# https://ru.stackoverflow.com/questions/320292/Проблема-с-заменой-слеша-в-пути-на-python


# txt_input работает, когда inf_input = False
# def prepare_input_filename(inf_input: bool, txt_input: bool, is_view: bool = False) -> (str, str):
#     """
#     Подготовка имени excel-файла к вводу данных из excel-файла
#     для функции the_xls_import
#     """
#     # Как с помощью модуля os получить название диска, в котором находится текущий каталог?
#     # https://ru.stackoverflow.com/questions/1079156/Как-с-помощью-модуля-os-получить-название-диска-в-котором-находится-текущий-кат
#     # Как проверить, существует ли файл при помощи Python
#     # https://python-scripts.com/file-exists
#     # Создание имени файла
#     # пока жестко фиксированное имя
#     drv_letter = pathlib.Path().absolute().drive
#     full_file_name = ''
#     curr_folder: str = os.path.abspath(os.curdir)
#     makroseis_fullpath = "\\".join([drv_letter, makroseis_folder])
#     if is_view:
#         print('curr_folder = ', curr_folder)
#     if is_view:
#         print('makroseis_fullpath = ', makroseis_fullpath)
#     if makroseis_fullpath == curr_folder:  # только тестирование
#         if inf_input:  # ввод полного комплекта с *.inf
#             test_filename = testinf_filename  # [если истина] if [выражение] else [если ложь]
#      #      res_file_name = test_restxt_filename if is_txt_res_file else test_resxls_filename
#         else:         # ввод только файла данных *.txt или *.xls
#             if txt_input:   # [если истина] if [выражение] else [если ложь]
#                 test_filename = testtxt_filename
#             else:
#                 test_filename = testxls_filename
#             res_ = test_restxt_filename if '_res.txt' else '_res.xlsx'
#             res_file_name = gfn(testxls_filename)+res_
#             print('res_file_name = ', res_file_name)
#     else: # работа
#         # print(ss_uc) # 'не сделано'
#         mb.showerror(s_error, ss_uc)
#         sys.exit()
#
#     full_file_name: str = "\\".join([curr_folder, datfolder, test_filename])
#     full_res_file_name =  "\\".join([curr_folder, datfolder, res_file_name])
#     path = pathlib.Path(full_file_name)
#     if not path.exists():
#         # print(ss_fpne, full_file_name) # 'путь не существует = '
#         mb.showerror(s_error, ss_fpne + full_file_name)
#     if not path.is_file():
#         # print('файл не существует = ', full_file_name) # 'файл не существует = '
#         mb.showerror(s_error, ss_ffne_ + full_file_name)
#
#     # if is_view: print('full_file_name = ', full_file_name)
#     return full_file_name, full_res_file_name
# # --------------- def prepare_datinput_filename()

def the_txt_importdat(txt_file_name: str,is_view: bool) -> object:
    """
    Чтение файла и  запись его в numpy.array
    """
    # http://pythonicway.com/python-fileio
    # numlines: int = text_file_num_lines(txt_file_name);
    # print('numlines = ',numlines)
    f = open(txt_file_name, 'r')
    # https://ru.stackoverflow.com/questions/928295/Чтение-файла-построчно-в-python
    # https://andreyex.ru/yazyk-programmirovaniya-python/uchebnik-po-python-3/python-3-strokovyj-metod-splitlines/
    ncol = 7
    all_lines = f.read().splitlines()  #  разделяет строку по символу переноса строки \n. Возвращает список(list)
    nrow1 = len(all_lines) # вместе со строкой заголовков
    ncol1 = ncol
    if is_view: print('ncol=', ncol)
    for i in range(nrow1):
        s: str = all_lines[i]
        all_lines[i] = s.strip()
        # print(i, all_lines[i])

    if is_view: print('Введено строк из текстового файла = ', nrow1)
    numpy_arr = np.zeros((nrow1 - 1, ncol), dtype=object)  # первую строку заголовков не вводим
    for i in range(nrow1):
        if i != 0:
            s = all_lines[i]            # print(s)
            part_lines = s.split(maxsplit=ncol1)    # print_string(part_lines)
            numpy_arr[i - 1, 0] = float(part_lines[0])  # Lat
            numpy_arr[i - 1, 1] = float(part_lines[1])  # Lon
            numpy_arr[i - 1, 2] = float(part_lines[2])  # Alt
            numpy_arr[i - 1, 3] = float(part_lines[3])  # I_fact
            numpy_arr[i - 1, 4] = float(part_lines[4])  # dI
            numpy_arr[i - 1, 5] = int(part_lines[5])    # N
            numpy_arr[i - 1, 6] = str(part_lines[6])    # Нас.пункт
            if len(part_lines) > ncol1:  # название состоит из 2 и более частей, в последней части остаток
                numpy_arr[i - 1, 6] = numpy_arr[i - 1, 6]+' '+str(part_lines[7])
    if is_view: view_2d_array(numpy_arr, nrow1 - 1, ncol1, '2d массив NumPy заполнен из строк')
    return numpy_arr
# --------------- def the_txt_import()


# Не используется, т.к. есть view_excel_sheet
def view_xlsx_info(my_wb):
    print('max_row = ', my_wb.max_row)
    print('max_col = ', my_wb.max_col)


def rebuild_xlsx(my_sheet, is_view: bool):
    # ==1== Удаление лишних столбцов
#  A      B     C     D     E     F     G     H   I     J  K   L      M          N        O       P            Q             R           S              T
#  1      2     3     4     5	  6	    7	  8   9    10  11  12     13        14        15      16           17            18          19             20
# Day STATION ALT.	GRAV.  SD.	TILTX TILTY	TEMP TIDE DUR REJ TIME DEC.TIME	Дата-Время	DATE "оср 20 мин" "лок 20 мин" "оср 2 часа" "лок 2 часа" "Дисп (20 мин)"
    # преобразование столбца 15 (O) только в дату и добавление туда точного времени из столбца 12
    for row_ in range(2,my_sheet.max_row+1): # 1 строка - заголовки, +1, чтобы пройти все строки
        cell_d1 = my_sheet.cell(row=row_, column=15);   sd1 = str(cell_d1.value);    sdo1 = sd1.split(sep=' ', maxsplit=2)
        cell_t1 = my_sheet.cell(row=row_, column=12);   st1 = str(cell_t1.value);    sto1 = st1.split(sep=' ', maxsplit=2)
        s = sdo1[0].strip() +' '+sto1[0].strip()
        cell_d1.value = datetime.datetime.strptime(s, '%Y-%m-%d %H:%M:%S')

    not_used_col = [20, 19, 18, 17, 16, 13, 1, 1] # 8
    for i in not_used_col:
        my_sheet.delete_cols(i)

    # ==2== Просмотр изменений
    if is_view: view_excel_sheet(my_sheet, 10)

    return my_sheet


def the_xls_importdat(xls_file_name: str, is_view: bool) -> object:
    """
    ВВод данных из excel-файла
    образец файл: Dat\точки_ввод.xlsx
    """
    nrow: int; ncol: int
#  2020 Чтение и запись в файл Excel с использованием модуля Python openpyxl
#  Источник: https://pythononline.ru/question/chtenie-i-zapis-v-fayl-excel-s-ispolzovaniem-modulya-python-openpyxl
    my_wb = openpyxl.load_workbook(xls_file_name, data_only=True)
    my_sheet = my_wb.active
    my_sheet_title = my_sheet.title
    if is_view:
        print("My sheet title: " + my_sheet_title)
# Openpyxl - Как найти количество строк с данными в xlsx
# Источник: https://question-it.com/questions/611944/openpyxl-kak-najti-kolichestvo-strok-s-dannymi-v-xlsx
# пустые столбцы и строки определяет плохо
#     if is_view: print('my_sheet.max_row = ', my_sheet.max_row)
#     if is_view: print('my_sheet.max_column = ',my_sheet.max_column)
    if is_view: view_excel_sheet(my_sheet, 40)
    # удаление лишних столбцов и создание нового с точным датой и временем
    my_sheet = rebuild_xlsx(my_sheet, is_view)
    # Подготовка массива NumPy
    # https://coderoad.ru/53659234/Как-записать-строку-в-массив-numpy
    nrow = my_sheet.max_row;    ncol = my_sheet.max_column
    # if is_view:
    #     print('nrow=',nrow)
    #     print('ncol=',ncol)
    #     input('press any key to continue')


    numpy_arr = np.zeros((nrow-1, ncol), dtype=object)  # первую строку заголовков не вводим
    # if is_view:   view_2d_array(numpy_arr, nrow-1, ncol, '2d массив NumPy создан')
    #  извлечение из my_sheet
    for i in range(nrow):
        if i != 0:  # первую строку заголовков не вводим
            for j in range(ncol):
                numpy_arr[i-1, j] = my_sheet.cell(row=i+1, column=j+1).value           # первую строку заголовков не вводим
    # if the_view:  view_2d_array(numpy_arr, nrow - 1, ncol, '2d массив NumPy заполнен')
    # if is_view:  view_2d_array(numpy_arr, nrow - 1, ncol, '2d массив NumPy заполнен')

    return numpy_arr
# --------------- def the_xls_import()


def view_excel_sheet(excel_sheet, n_view_row=-1) -> None:
    print('2d лист excel')
    nrow: int = excel_sheet.max_row
    ncol: int = excel_sheet.max_column
    print('excel_sheet.max_row = ', nrow)
    print('excel_sheet.max_column = ', ncol)

    n = n_view_row if n_view_row>0 else nrow
    shift_str = repeat_chr(' ',9)
    for i in range(n):
        # Использование Python и Excel для обработки и анализа данных. Часть 2: библиотеки для работы с данными
        # https://habr.com/ru/company/otus/blog/331998/
        if i != 0: # номера значений
            print(format(i, '7d'), end='')
        else:
            print(shift_str, end='')
        for j in range(ncol):
            s = shift_str if j == 0 else ''  # сдвиг из-за номера значений
            s = s + str(excel_sheet.cell(row=i+1, column=j+1).value)
            lens = len(s)
            if lens<=9: the_s = '9s'
            else: the_s = str(lens)+'s'
            print(format(s,the_s), end='  ')
        print()



def view_2d_array(arr, nrow: int, ncol: int, test_info: str = '') -> None:
    """
    Вывод на печать 2 мерного массива
    """
    print(test_info)
    # input('input nn')
    for i in range(nrow):
        # Функция format() в Python, форматирует значение переменной
        # https://docs-python.ru/tutorial/vstroennye-funktsii-interpretatora-python/funktsija-format/
        print(format(i, '3d'), end=' ')
        for j in range(ncol):
            print(arr[i, j], end=' ')
        print()


def view_type2d_array(arr, nrow: int, ncol: int, test_info: str='') -> None:
    """
    Вывод на печать 2 мерного массива
    """
    print(test_info)
    for i in range(nrow):
        # Функция format() в Python, форматирует значение переменной
        # https://docs-python.ru/tutorial/vstroennye-funktsii-interpretatora-python/funktsija-format/
        print(format(i, '3d'), end=' ')
        for j in range(ncol):
            print(type(arr[i, j]), end=' ')
        print()


def the_input_dat(file_dat_name: str, is_view: bool = False) -> (object, str):
    """
    Общая функция ввода
    Как получить дату и время создания файла в Python
    https://issue.life/questions/237079
    https://docs.python.org/3/library/os.path.html#os.path.getmtime
    """
    # filename: str = prepare_input_datfilename(txt_input)
    # print(filename)
    # print(get_file_time(filename))
    matrix_file_name =''
    ext: str = gfe(file_dat_name)  # расширение с точкой в нижнем регистре
    if ext == '.txt':
        numpy_arr = the_txt_importdat(file_dat_name, is_view)
    else:
        # dt_now = datetime.datetime.now();   print(dt_now)
        numpy_arr = the_xls_importdat(file_dat_name, is_view=False)  # ввели данные
        # dt_now = datetime.datetime.now();   print(dt_now)
        matrix_file_name = create_matrix_fn(file_dat_name)
        # dt_now = datetime.datetime.now();   print(dt_now)
        save_in_matrix(matrix_file_name, numpy_arr)
        matrix_file_name = matrix_file_name + '.npy'
        # dt_now = datetime.datetime.now();   print(dt_now)

    return (numpy_arr, matrix_file_name)


def create_matrix_fn(file_dat_name:str)->str:
    s = pfile.gfpnwe(file_dat_name)+'_mat'
    return s

def create_stat_fn(file_dat_name:str)->str:
    s = pfile.gfpnwe(file_dat_name)+'_stat.txt'
    return s


def save_in_matrix(fn:str, numpy_arr):
    """
    сохранение в файл для убыстрения повторного считывания
    6. Файловый ввод и вывод массивов
    https://pyprog.pro/file_input_and_output.html
    """
    np.save(fn,numpy_arr,allow_pickle=True)
    return fn

def load_from_matrix(fn:str):
    """
    чтение из файла для убыстрения повторных операций
    6. Файловый ввод и вывод массивов
    https://pyprog.pro/file_input_and_output.html
    """
    numpy_arr = np.load(fn,allow_pickle=True)
    return numpy_arr

def get_file_time(fname: str, thetest: bool = False) -> str:
    """
    Вывод времени создания файла
    2018 Python 3 – Время. Метод strftime()
    https://andreyex.ru/yazyk-programmirovaniya-python/uchebnik-po-python-3/python-3-vremya-metod-strftime/
    """
    beauty_time: float = os.path.getmtime(fname)
    if thetest: print(os.path.getmtime(fname))
    beauty_time1: str = time.strftime("%d/%m/%y : %H:%M:%S", time.localtime(beauty_time))  # gmtime
    print('Время файла = ', beauty_time1)
    if thetest: print('type(beauty_time) = ', type(beauty_time1))
    return beauty_time1


def the_input(fname: str, res_dir: str, is_view: bool, matrix_file_name:str) -> (bool, object):
    """
    """
    (good_end, thecurr_dict) = input_inf(fname, is_view)
    if not good_end:
        mb.showerror(s_error, ss_fifnf)  # 'inf - файл не найден'
        return False, None
    else:
        if not control_curr_dict(thecurr_dict):
            mb.showerror(s_error, ss_feif)  # 'Ошибки в inf-файле'
            return False, None
        else:
            fdat_name = name_and_ext(thecurr_dict["work_dir"], thecurr_dict["fdat_name_"])
            if is_view: print('fdat_name = ', fdat_name)
            if matrix_file_name =='':
                (numpy_arr, matrix_file_name) = the_input_dat(fdat_name, is_view)
                thecurr_dict["matrix_file_name"] = matrix_file_name
            else:
                numpy_arr = load_from_matrix(matrix_file_name)
# ndarray.shape - размеры массива, его форма. Это кортеж натуральных чисел, показывающий длину массива по каждой оси.
# Для матрицы из n строк и m столбов, shape будет (n,m).
            (row, col) = np.shape(numpy_arr)
            if is_view: print(row, col)
            thecurr_dict["npoint"] = row
            add_dat_struct(thecurr_dict, numpy_arr)
            if is_view:
                print('print_dat_struct()')
                print_dat_struct()
            pmain_proc.log_file_name = create_log_filename(res_dir, fname)
            return True, thecurr_dict


def prc(row: int, col: int) -> None:  # для проверки печать строки и столбца
    print('   row = ', row)
    print('column = ', col)


def create_log_filename(resdir: str, inf_file_name) -> str:
    fn = gfn(inf_file_name) + '_res.xlsx'
    s = "\\".join([resdir, fn])
    return s
